# api/investor.py
from app import investor_crud
from fastapi import APIRouter, Depends, HTTPException, status,Query, File, UploadFile,Request
from fastapi.security import OAuth2PasswordBearer
from starlette.responses import FileResponse
from sqlalchemy.orm import Session
from sqlalchemy.exc import SQLAlchemyError
from fastapi.responses import JSONResponse
from fastapi import Response
from app.deps import get_db
from datetime import datetime,date
from reportlab.pdfgen import canvas
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from app.schemas import investor_schemas
from app.models import investor_models
from typing import List
import uuid
import shutil
import os
import io
import logging

from app.schemas.investor_schemas import UnclaimedAmountResponse,SwitchCreateSchema,SwitchResponseSchema,SchemeResponse,FolioSummaryResponse,SipResponse
from app.schemas.investor_schemas import CapitalGainRecord ,PurchaseRequest,RedemptionRequest, TransactionLedgerResponse,SIPRegistrationCreate,SIPRegistrationResponse,STPCreateSchema 
from app.schemas.investor_schemas import SessionsResponse,TwoFAStatusResponse
from app.api.investor_auth import get_current_investor
from app.models.investor_models import FolioSummary, PurchaseTransaction,RedemptionTransaction, TransactionLedger, Investor, SwitchTransaction,SIPTransaction
from app.util import generate_unique_folio_number

router = APIRouter()
oauth2_scheme = OAuth2PasswordBearer(tokenUrl="/auth/login")

@router.get("/profile", response_model=investor_schemas.InvestorResponse)
def get_profile(current_investor: investor_models.Investor = Depends(get_current_investor)):
    return current_investor

@router.put("/profile", response_model=investor_schemas.InvestorResponse)
def update_profile(
    updated_data: investor_schemas.InvestorUpdate,
    db: Session = Depends(get_db),
    current_investor: investor_models.Investor = Depends(get_current_investor),
):
    # Only update allowed fields
    for field, value in updated_data.dict(exclude_unset=True).items():
        setattr(current_investor, field, value)
    db.commit()
    db.refresh(current_investor)
    return current_investor

@router.post("/banks", response_model=investor_schemas.BankResponse)
def add_bank(
    bank: investor_schemas.BankCreate,
    db: Session = Depends(get_db),
    current_investor: investor_models.Investor = Depends(get_current_investor),
):
    db_bank = investor_models.Bank(
        investor_id=current_investor.investor_id,
        account_no=bank.account_no,
        ifsc=bank.ifsc,
        branch=bank.branch,
        
    )
    db.add(db_bank)
    db.commit()
    db.refresh(db_bank)
    return db_bank

@router.put("/banks/{bank_id}", response_model=investor_schemas.BankResponse)
def update_bank(
    bank_id: int,
    bank_update: investor_schemas.BankCreate,
    db: Session = Depends(get_db),
    current_investor: investor_models.Investor = Depends(get_current_investor),
):
    db_bank = (
        db.query(investor_models.Bank)
        .filter(investor_models.Bank.bank_id == bank_id, investor_models.Bank.investor_id == current_investor.investor_id)
        .first()
    )
    if not db_bank:
        raise HTTPException(status_code=404, detail="Bank not found")
    for field, value in bank_update.dict(exclude_unset=True).items():
        setattr(db_bank, field, value)
    db.commit()
    db.refresh(db_bank)
    return db_bank

@router.delete("/banks/{bank_id}", status_code=204)
def delete_bank(
    bank_id: int,
    db: Session = Depends(get_db),
    current_investor: investor_models.Investor = Depends(get_current_investor),
):
    db_bank = (
        db.query(investor_models.Bank)
        .filter(investor_models.Bank.bank_id == bank_id, investor_models.Bank.investor_id == current_investor.investor_id)
        .first()
    )
    if not db_bank:
        raise HTTPException(status_code=404, detail="Bank not found")
    db.delete(db_bank)
    db.commit()
    return

@router.post("/nominees", response_model=investor_schemas.NomineeResponse)
def add_nominee(
    nominee: investor_schemas.NomineeCreate,
    db: Session = Depends(get_db),
    current_investor: investor_models.Investor = Depends(get_current_investor),
):
    db_nominee = investor_models.Nominee(
        investor_id=current_investor.investor_id,
        name=nominee.name,
        relation=nominee.relation,
        pct=nominee.pct,
    )
    db.add(db_nominee)
    db.commit()
    db.refresh(db_nominee)
    return db_nominee


@router.put("/nominees/{nominee_id}", response_model=investor_schemas.NomineeResponse)
def update_nominee(
    nominee_id: int,
    nominee_update: investor_schemas.NomineeCreate,
    db: Session = Depends(get_db),
    current_investor: investor_models.Investor = Depends(get_current_investor),
):
    db_nominee = (
        db.query(models.Nominee)
        .filter(investor_models.Nominee.nominee_id == nominee_id, investor_models.Nominee.investor_id == current_investor.investor_id)
        .first()
    )
    if not db_nominee:
        raise HTTPException(status_code=404, detail="Nominee not found")
    for field, value in nominee_update.dict(exclude_unset=True).items():
        setattr(db_nominee, field, value)
    db.commit()
    db.refresh(db_nominee)
    return db_nominee

@router.delete("/nominees/{nominee_id}", status_code=204)
def delete_nominee(
    nominee_id: int,
    db: Session = Depends(get_db),
    current_investor: investor_models.Investor = Depends(get_current_investor),
):
    db_nominee = (
        db.query(investor_models.Nominee)
        .filter(investor_models.Nominee.nominee_id == nominee_id, investor_models.Nominee.investor_id == current_investor.investor_id)
        .first()
    )
    if not db_nominee:
        raise HTTPException(status_code=404, detail="Nominee not found")
    db.delete(db_nominee)
    db.commit()
    return


@router.post("/purchase", response_model=TransactionLedgerResponse, status_code=201)
def make_purchase(
    purchase: PurchaseRequest,
    db: Session = Depends(get_db),
    current_investor: Investor = Depends(get_current_investor),
):
    nav = 250.0  # Replace with actual NAV fetch logic
    units = purchase.amount / nav

    # Fetch or create folio holding
    folio = db.query(FolioSummary).filter(
        FolioSummary.investor_id == current_investor.investor_id,
        FolioSummary.scheme_id == purchase.scheme_id,
    ).first()

    if folio:
        folio.total_units += units
        folio.total_value += purchase.amount
        folio.last_nav = nav
        folio.updated_at = datetime.utcnow()
    else:
        new_folio_number = generate_unique_folio_number(current_investor.investor_id, purchase.scheme_id)
        folio = FolioSummary(
            folio_number=new_folio_number,
            investor_id=current_investor.investor_id,
            scheme_id=purchase.scheme_id,
            total_units=units,
            total_value=purchase.amount,
            last_nav=nav,
            updated_at=datetime.utcnow()
        )
        db.add(folio)
        db.commit()
        db.refresh(folio)

    # Insert into PurchaseTransaction table
    purchase_txn = PurchaseTransaction(
        investor_id=current_investor.investor_id,
        folio_number=folio.folio_number,
        scheme_id=purchase.scheme_id,
        amount=purchase.amount,
        nav=nav,
        units=units,
        status="Pending Payment",
        date=datetime.utcnow(),
        bank=purchase.bank,
        payment_mode=purchase.payment_mode,
    )
    db.add(purchase_txn)
    db.commit()
    db.refresh(purchase_txn)

    # Insert into Transaction ledger table
    ledger_txn = TransactionLedger(
        investor_id=current_investor.investor_id,
        folio_number=folio.folio_number,
        txn_type="Fresh Purchase",
        scheme_id=purchase.scheme_id,
        amount=purchase.amount,
        nav=nav,
        units=units,
        status="Pending Payment",
        date=datetime.utcnow(),
        bank=purchase.bank,
        payment_mode=purchase.payment_mode,
    )
    db.add(ledger_txn)
    db.commit()
    db.refresh(ledger_txn)

    return ledger_txn

@router.post("/redemption", response_model=TransactionLedgerResponse, status_code=201)
def make_redemption(
    redemption: RedemptionRequest,
    db: Session = Depends(get_db),
    current_investor: Investor = Depends(get_current_investor),
):
    folio = db.query(FolioSummary).filter(
        FolioSummary.scheme_id == redemption.scheme_id,
        FolioSummary.investor_id == current_investor.investor_id,
    ).first()

    if not folio:
        raise HTTPException(status_code=404, detail="Folio not found for this scheme")

    # Decide units to redeem based on request
    if redemption.amount:
        units_to_redeem = redemption.amount / folio.last_nav
    elif redemption.units:
        units_to_redeem = redemption.units
    else:
        units_to_redeem = folio.total_units  # Full redemption

    if units_to_redeem > folio.total_units:
        raise HTTPException(status_code=400, detail="Redemption units exceed holding")

    amount_to_redeem = units_to_redeem * folio.last_nav

    # Insert RedemptionTransaction
    redemption_txn = RedemptionTransaction(
        investor_id=current_investor.investor_id,
        folio_number=folio.folio_number,
        scheme_id=redemption.scheme_id,
        amount=amount_to_redeem,
        units=units_to_redeem,
        nav=folio.last_nav,
        date=datetime.utcnow(),
        status="Pending",
        bank=redemption.bank,
        payment_mode=redemption.payment_mode,
        txn_type="Redemption",
    )
    db.add(redemption_txn)

    # Insert into TransactionLedger
    ledger_txn = TransactionLedger(
        investor_id=current_investor.investor_id,
        folio_number=folio.folio_number,
        txn_type="Redemption",
        scheme_id=redemption.scheme_id,
        amount=amount_to_redeem,
        units=units_to_redeem,
        nav=folio.last_nav,
        date=datetime.utcnow(),
        status="Pending",
        bank=redemption.bank,
        payment_mode=redemption.payment_mode,
    )
    db.add(ledger_txn)

    # Update folio summary units and value
    folio.total_units -= units_to_redeem
    folio.total_value = folio.total_units * folio.last_nav
    folio.updated_at = datetime.utcnow()

    db.commit()
    db.refresh(redemption_txn)

    return redemption_txn


@router.post("/sip", response_model=investor_schemas.SIPRegistrationResponse, status_code=201)
def create_sip(
    sip: investor_schemas.SIPRegistrationCreate,
    db: Session = Depends(get_db),
    current_investor: investor_models.Investor = Depends(get_current_investor)
):
    # Check if folio provided, else try to find existing folio or create new
    if sip.folio_number:
        folio_number = sip.folio_number
    else:
        folio = db.query(investor_models.FolioSummary).filter(
            investor_models.FolioSummary.investor_id == current_investor.investor_id,
            investor_models.FolioSummary.scheme_id == sip.scheme_id,
        ).first()
        if folio:
            folio_number = folio.folio_number
        else:
            folio_number = generate_unique_folio_number(current_investor.investor_id, sip.scheme_id)
            folio = investor_models.FolioSummary(
                folio_number=folio_number,
                investor_id=current_investor.investor_id,
                scheme_id=sip.scheme_id,
                total_units=0,
                total_value=0,
                last_nav=0,
                updated_at=datetime.utcnow()
            )
            db.add(folio)
            db.commit()
            db.refresh(folio)

    reg_id = "SIP" + str(uuid.uuid4())[:8]

    # Create SIP registration record
    db_sip = investor_models.SIPTransaction(
        reg_id=reg_id,
        investor_id=current_investor.investor_id,
        folio_number=folio_number,
        scheme_id=sip.scheme_id,
        amount=sip.amount,
        frequency=sip.frequency,
        start_date=sip.start_date,
        end_date=sip.end_date,
        installments=sip.installments,
        bank=sip.bank,
        mandate=sip.mandate,
        status="Active",
        date=datetime.utcnow()
    )
    db.add(db_sip)

    # Calculate units based on NAV (adjust NAV fetching logic as required)
    nav = 250.0
    units = sip.amount / nav

    # Create a transaction ledger entry
    txn = investor_models.TransactionLedger(
        investor_id=current_investor.investor_id,
        folio_number=folio_number,
        txn_type="SIP Setup",
        scheme_id=sip.scheme_id,
        amount=sip.amount,
        nav=nav,
        units=units,
        status="Pending Payment",
        bank=sip.bank,
        payment_mode=sip.mandate,
        date=datetime.utcnow()
    )
    db.add(txn)

    # Update folio summary units and value
    folio.total_units += units
    folio.total_value += sip.amount
    folio.last_nav = nav
    folio.updated_at = datetime.utcnow()

    db.commit()
    db.refresh(db_sip)

    return db_sip

@router.get("/sip/active", response_model=List[investor_schemas.SIPRegistrationResponse])
def get_active_sips(
    db: Session = Depends(get_db),
    current_investor: investor_models.Investor = Depends(get_current_investor)
):
    sips = db.query(investor_models.SIPTransaction).filter(
        investor_models.SIPTransaction.investor_id == current_investor.investor_id,
        investor_models.SIPTransaction.status == "Active"
    ).all()
    return sips

@router.post("/swp", response_model=investor_schemas.SWPRegistrationResponse, status_code=201)
def create_swp(
    swp: investor_schemas.SWPRegistrationCreate,
    db: Session = Depends(get_db),
    current_investor: investor_models.Investor = Depends(get_current_investor),
):
    try:
        if swp.folio_number:
            folio_number = swp.folio_number
        else:
            folio = db.query(investor_models.FolioSummary).filter(
                investor_models.FolioSummary.investor_id == current_investor.investor_id,
                investor_models.FolioSummary.scheme_id == swp.scheme_id,
            ).first()
            if folio:
                folio_number = folio.folio_number
            else:
                folio_number = generate_unique_folio_number(current_investor.investor_id, swp.scheme_id)
                folio = investor_models.FolioSummary(
                    folio_number=folio_number,
                    investor_id=current_investor.investor_id,
                    scheme_id=swp.scheme_id,
                    total_units=0,
                    total_value=0,
                    last_nav=0,
                    updated_at=datetime.utcnow()
                )
                db.add(folio)
                db.commit()
                db.refresh(folio)

        nav_value = folio.last_nav or 250.0
        if swp.units is not None:
            units_to_redeem = swp.units
            amount_to_redeem = units_to_redeem * nav_value
        else:
            amount_to_redeem = swp.amount
            units_to_redeem = amount_to_redeem / nav_value if nav_value else 0

        if units_to_redeem > folio.total_units:
            raise HTTPException(status_code=400, detail="SWP units exceed holding")

        reg_id = "SWP" + str(uuid.uuid4())[:8]
        db_swp = investor_models.SWPTransaction(
            reg_id=reg_id,
            investor_id=current_investor.investor_id,
            folio_number=folio_number,
            scheme_id=swp.scheme_id,
            amount=amount_to_redeem,
            units=units_to_redeem,
            nav=nav_value,
            frequency=swp.frequency,
            date=datetime.utcnow(),
            status="Pending",
            bank=swp.bank,
            payment_mode=swp.mandate
        )
        db.add(db_swp)

        txn = investor_models.TransactionLedger(
            investor_id=current_investor.investor_id,
            folio_number=folio_number,
            txn_type="SWP",
            scheme_id=swp.scheme_id,
            amount=amount_to_redeem,
            units=units_to_redeem,
            nav=nav_value,
            status="Pending",
            bank=swp.bank,
            payment_mode=swp.mandate,
            date=datetime.utcnow()
        )
        db.add(txn)

        folio.total_units -= units_to_redeem
        folio.total_value = folio.total_units * nav_value
        folio.updated_at = datetime.utcnow()

        db.commit()
        db.refresh(db_swp)

        return db_swp

    except SQLAlchemyError as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Database error: {str(e)}")
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Server error: {str(e)}")

@router.post("/stp", status_code=201)
def create_stp(
    stp: investor_schemas.STPCreateSchema,
    db: Session = Depends(get_db),
    current_investor: investor_models.Investor = Depends(get_current_investor),
):
    try:
        source_folio = db.query(investor_models.FolioSummary).filter(
            investor_models.FolioSummary.investor_id == current_investor.investor_id,
            investor_models.FolioSummary.scheme_id == stp.from_scheme_id,
        ).first()

        if not source_folio or source_folio.total_units <= 0:
            raise HTTPException(status_code=400, detail="Invalid folio or insufficient units")

        nav_from = source_folio.last_nav or 1.0

        if stp.units is not None:
            units_to_redeem = stp.units
            amount_to_redeem = units_to_redeem * nav_from
        else:
            amount_to_redeem = stp.amount
            units_to_redeem = amount_to_redeem / nav_from

        if units_to_redeem > source_folio.total_units:
            raise HTTPException(status_code=400, detail="Units exceed holding")

        if amount_to_redeem > source_folio.total_value:
            raise HTTPException(status_code=400, detail="Amount exceeds folio value")

        source_folio.total_units -= units_to_redeem
        source_folio.total_value -= amount_to_redeem
        source_folio.updated_at = datetime.utcnow()
        db.add(source_folio)

        redemption_txn = investor_models.TransactionLedger(
            investor_id=current_investor.investor_id,
            folio_number=source_folio.folio_number,
            txn_type="STP Redemption",
            scheme_id=stp.from_scheme_id,
            amount=amount_to_redeem,
            nav=nav_from,
            units=-units_to_redeem,
            status="Pending Payment",
            date=stp.date,
        )
        db.add(redemption_txn)

        dest_folio = db.query(investor_models.FolioSummary).filter(
            investor_models.FolioSummary.investor_id == current_investor.investor_id,
            investor_models.FolioSummary.scheme_id == stp.to_scheme_id,
        ).first()

        if not dest_folio:
            # Implement your unique folio number generator appropriately
            dest_folio = investor_models.FolioSummary(
                folio_number=generate_unique_folio_number(current_investor.investor_id, stp.to_scheme_id),
                investor_id=current_investor.investor_id,
                scheme_id=stp.to_scheme_id,
                total_units=0,
                total_value=0,
                last_nav=0,
                updated_at=datetime.utcnow(),
            )
            db.add(dest_folio)

        nav_to = dest_folio.last_nav or 1.0
        units_to_purchase = amount_to_redeem / nav_to
        dest_folio.total_units += units_to_purchase
        dest_folio.total_value += amount_to_redeem
        dest_folio.updated_at = datetime.utcnow()
        db.add(dest_folio)

        purchase_txn = investor_models.TransactionLedger(
            investor_id=current_investor.investor_id,
            folio_number=dest_folio.folio_number,
            txn_type="STP Purchase",
            scheme_id=stp.to_scheme_id,
            amount=amount_to_redeem,
            nav=nav_to,
            units=units_to_purchase,
            status="Pending Payment",
            date=stp.date,
        )
        db.add(purchase_txn)
        stp_txn = investor_models.STPTransaction(
        investor_id=current_investor.investor_id,
        folio_number=source_folio.folio_number,
        from_scheme_id=stp.from_scheme_id,
        to_scheme_id=stp.to_scheme_id,
        amount=amount_to_redeem,
        units=units_to_redeem,
        nav=nav_from,
        date=stp.date,
        status="Pending"
        )
        db.add(stp_txn)

        db.commit()

        return {
            "redemption_txn_id": redemption_txn.txn_id,
            "purchase_txn_id": purchase_txn.txn_id,
            "message": "STP transactions created successfully",
        }
    except SQLAlchemyError as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Database error: {str(e)}")
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Server error: {str(e)}")


@router.post("/switch", status_code=201, response_model=SwitchResponseSchema)
def create_switch(
    switch_req: SwitchCreateSchema,
    db: Session = Depends(get_db),
    current_investor: Investor = Depends(get_current_investor),
):
    # Fetch source folio summary for the from_scheme_id
    source_folio = db.query(FolioSummary).filter(
        FolioSummary.investor_id == current_investor.investor_id,
        FolioSummary.scheme_id == switch_req.from_scheme_id,
    ).first()

    if not source_folio or source_folio.total_units <= 0:
        raise HTTPException(status_code=400, detail="Invalid folio or insufficient units")

    nav_from = source_folio.last_nav or 1

    # Calculate amount and units (treating interchangeably)
    if switch_req.units:
        units = switch_req.units
        amount = units * nav_from
    else:
        amount = switch_req.amount
        units = amount / nav_from

    # Validate source folio holdings sufficiency
    if units > source_folio.total_units:
        raise HTTPException(status_code=400, detail="Units exceed source folio holdings")
    if amount > source_folio.total_value:
        raise HTTPException(status_code=400, detail="Amount exceeds source folio value")

    # Fetch or create destination folio
    dest_folio = db.query(FolioSummary).filter(
        FolioSummary.investor_id == current_investor.investor_id,
        FolioSummary.scheme_id == switch_req.to_scheme_id,
    ).first()

    if not dest_folio:
        folio_num = generate_unique_folio_number(current_investor.investor_id, switch_req.to_scheme_id)
        dest_folio = FolioSummary(
            folio_number=folio_num,
            investor_id=current_investor.investor_id,
            scheme_id=switch_req.to_scheme_id,
            total_units=0,
            total_value=0,
            last_nav=1,
            updated_at=datetime.utcnow(),
        )
        db.add(dest_folio)

    nav_to = dest_folio.last_nav or 1
    units_to_purchase = amount / nav_to

    # Update source and dest folios
    source_folio.total_units -= units
    source_folio.total_value -= amount
    source_folio.updated_at = datetime.utcnow()

    dest_folio.total_units += units_to_purchase
    dest_folio.total_value += amount
    dest_folio.updated_at = datetime.utcnow()

    db.add(source_folio)
    db.add(dest_folio)

    # Create transactions in transaction ledger
    redemption_txn = TransactionLedger(
        investor_id=current_investor.investor_id,
        folio_number=source_folio.folio_number,
        txn_type="Switch Redemption",
        scheme_id=switch_req.from_scheme_id,
        amount=amount,
        nav=nav_from,
        units=-units,
        status="Pending",
        date=switch_req.date or datetime.utcnow(),
    )
    purchase_txn = TransactionLedger(
        investor_id=current_investor.investor_id,
        folio_number=dest_folio.folio_number,
        txn_type="Switch Purchase",
        scheme_id=switch_req.to_scheme_id,
        amount=amount,
        nav=nav_to,
        units=units_to_purchase,
        status="Pending",
        date=switch_req.date or datetime.utcnow(),
    )
    db.add(redemption_txn)
    db.add(purchase_txn)

    # Record switch transaction
    switch_txn = SwitchTransaction(
        investor_id=current_investor.investor_id,
        folio_number=source_folio.folio_number,
        from_scheme_id=switch_req.from_scheme_id,
        to_scheme_id=switch_req.to_scheme_id,
        amount=amount,
        units=units,
        nav=nav_from,
        date=switch_req.date or datetime.utcnow(),
        status="Pending",
    )
    db.add(switch_txn)

    db.commit()

    return SwitchResponseSchema(
        redemption_txn_id=redemption_txn.txn_id,
        purchase_txn_id=purchase_txn.txn_id,
        switch_txn_id=switch_txn.txn_id,
        message="Switch transactions created successfully",
    )



@router.get("/investor/schemes", response_model=List[SchemeResponse])
def get_investor_schemes(
    db: Session = Depends(get_db),
    current_investor=Depends(get_current_investor),
):
    schemes = (
        db.query(FolioSummary.scheme_id)
        .filter(FolioSummary.investor_id == current_investor.investor_id)
        .distinct()
        .all()
    )
    # schemes is a list of tuples, convert to list of dicts
    return [{"scheme_id": s[0]} for s in schemes]

@router.get("/investor/transactionledger", response_model=List[TransactionLedgerResponse])
def get_investor_transactions(
    db: Session = Depends(get_db),
    current_investor=Depends(get_current_investor),
):
    # Query all transactions for the current investor
    transactions = (
        db.query(TransactionLedger)
        .filter(TransactionLedger.investor_id == current_investor.investor_id)
        .order_by(TransactionLedger.date.desc()) # recent first
        .all()
    )
    return transactions

@router.get("/investor/foliosummary", response_model=List[FolioSummaryResponse])
def get_investor_folios(
    db: Session = Depends(get_db),
    current_investor=Depends(get_current_investor),
):
    folios = (
        db.query(FolioSummary)
        .filter(FolioSummary.investor_id == current_investor.investor_id)
        .all()
    )
    return folios

@router.get("/investor/sips", response_model=List[SipResponse])
def get_active_sips(
    db: Session = Depends(get_db),
    current_investor=Depends(get_current_investor),
):
    sips = (
        db.query(SIPTransaction)
        .filter(SIPTransaction.investor_id == current_investor.investor_id)
        .all()
    )
    return sips

@router.get("/idcw-preferences", response_model=List[investor_schemas.IDCWPreferenceResponse])
def get_idcw_preferences(
    db: Session = Depends(get_db),
    current_investor: investor_models.Investor = Depends(get_current_investor),
):
    prefs = db.query(investor_models.IDCWPreference).filter(investor_models.IDCWPreference.investor_id == current_investor.investor_id).all()
    return prefs


@router.post("/idcw-preferences")
def set_idcw_preferences(
    preferences: List[investor_schemas.IDCWPreferenceCreate],
    db: Session = Depends(get_db),
    current_investor: investor_models.Investor = Depends(get_current_investor),
):
    for pref in preferences:
        existing = (
            db.query(investor_models.IDCWPreference)
            .filter(
                investor_models.IDCWPreference.investor_id == current_investor.investor_id,
                investor_models.IDCWPreference.scheme_id == pref.scheme_id,
            )
            .first()
        )
        if existing:
            existing.preference = pref.preference
        else:
            new_pref = investor_models.IDCWPreference(
                investor_id=current_investor.investor_id,
                scheme_id=pref.scheme_id,
                preference=pref.preference,
            )
            db.add(new_pref)
    db.commit()
    return {"message": "IDCW preferences updated successfully"}



@router.get("/investor/unclaimed-amounts", response_model=List[investor_schemas.UnclaimedAmountResponse])
def get_unclaimed_amounts(
    db: Session = Depends(get_db),
    current_investor: investor_models.Investor = Depends(get_current_investor),
):
    unclaimed_records = (
        db.query(investor_models.UnclaimedAmount)
        .join(investor_models.TransactionLedger, investor_models.UnclaimedAmount.txn_id == investor_models.TransactionLedger.txn_id)
        .filter(investor_models.UnclaimedAmount.investor_id == current_investor.investor_id)
        .all()
    )

    results = []
    for record in unclaimed_records:
        txn = record.transaction
        scheme_name = txn.scheme_id if txn else ""
        folio_number = txn.folio_number if txn else ""
        txn_type = txn.txn_type if txn else ""
        desc = "Automatically flagged due to uncredited payment"
        results.append(
            investor_schemas.UnclaimedAmountResponse(
                unclaimed_id=record.unclaimed_id,
                txn_id=txn.txn_id if txn else 0,
                scheme=scheme_name,
                folio=folio_number,
                type=txn_type,
                amount=record.amount,
                unclaimed_since=record.start_date,
                status=record.status,
                description=desc,
            )
        )
    return results


@router.post("/investor/unclaimed-claim")
def claim_unclaimed_amount(
    unclaimed_id: int,
    db: Session = Depends(get_db),
    current_investor: investor_models.Investor = Depends(get_current_investor),
):
    record = db.query(investor_models.UnclaimedAmount).filter(
        investor_models.UnclaimedAmount.unclaimed_id == unclaimed_id,
        investor_models.UnclaimedAmount.investor_id == current_investor.investor_id,
    ).first()
    if not record:
        raise HTTPException(status_code=404, detail="Unclaimed amount not found")
    if record.status == "Claimed":
        raise HTTPException(status_code=400, detail="Already claimed")

    record.status = "Claimed"
    db.commit()
    return {"message": f"Unclaimed amount {unclaimed_id} claimed successfully"}


def sync_pending_redemptions_to_unclaimed(db: Session):
    # Query pending Redemption/SWP transactions
    pending_txns = (
        db.query(investor_models.TransactionLedger)
        .filter(
            investor_models.TransactionLedger.txn_type.in_(["Redemption", "SWP"]),
            investor_models.TransactionLedger.status == "Pending"
        )
        .all()
    )

    for txn in pending_txns:
        # Skip if already recorded in unclaimed_amounts
        existing = (
            db.query(investor_models.UnclaimedAmount)
            .filter(investor_models.UnclaimedAmount.txn_id == txn.txn_id)
            .first()
        )
        if existing:
            continue

        # Create new unclaimed amount record
        new_unclaimed = investor_models.UnclaimedAmount(
            investor_id=txn.investor_id,
            txn_id=txn.txn_id,
            amount=txn.amount,
            status="Pending",
            start_date=date.today(),
        )
        db.add(new_unclaimed)

    db.commit()

@router.post("/investor/sync-unclaimed-amounts")
def sync_unclaimed_amounts_route(db: Session = Depends(get_db)):
    try:
        sync_pending_redemptions_to_unclaimed(db)
        return {"message": "Unclaimed amounts synced successfully"}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


def calculate_and_store_capital_gains(db: Session, investor_id: int, start_date: date, end_date: date):
    sell_txns = (
        db.query(investor_models.TransactionLedger)
        .filter(
            investor_models.TransactionLedger.investor_id == investor_id,
            investor_models.TransactionLedger.txn_type.in_(["Redemption", "Sale"]),  # Adjust types as needed
            investor_models.TransactionLedger.date >= start_date,
            investor_models.TransactionLedger.date <= end_date,
            investor_models.TransactionLedger.status == "Completed",
        )
        .all()
    )

    for sell_txn in sell_txns:
        buy_txn = (
            db.query(investor_models.TransactionLedger)
            .filter(
                investor_models.TransactionLedger.folio_number == sell_txn.folio_number,
                investor_models.TransactionLedger.scheme_id == sell_txn.scheme_id,
                investor_models.TransactionLedger.txn_type.in_(["Fresh Purchase", "Additional Purchase"]),
                investor_models.TransactionLedger.date < sell_txn.date,
                investor_models.TransactionLedger.status == "Completed",
                investor_models.TransactionLedger.investor_id == investor_id,
            )
            .order_by(investor_models.TransactionLedger.date.asc())
            .first()
        )
        if buy_txn:
            gains = sell_txn.amount - buy_txn.amount
            long_term = (sell_txn.date - buy_txn.date).days >= 365
            gain_type = "Long-term" if long_term else "Short-term"

            existing = (
                db.query(investor_models.CapitalGains)
                .filter(
                    investor_models.CapitalGains.investor_id == investor_id,
                    investor_models.CapitalGains.buy_date == buy_txn.date.date(),
                    investor_models.CapitalGains.sell_date == sell_txn.date.date(),
                    investor_models.CapitalGains.folio_number == sell_txn.folio_number,
                    investor_models.CapitalGains.scheme_id == sell_txn.scheme_id,
                )
                .first()
            )
            if existing:
                continue

            new_gain = investor_models.CapitalGains(
                investor_id=investor_id,
                scheme_id=sell_txn.scheme_id,
                folio_number=sell_txn.folio_number,
                buy_date=buy_txn.date.date(),
                sell_date=sell_txn.date.date(),
                buy_amount=buy_txn.amount,
                sell_amount=sell_txn.amount,
                gains=gains,
                gain_type=gain_type,
            )
            db.add(new_gain)

    db.commit()


@router.post("/investor/capital-gains-sync")
def sync_capital_gains(
    year: str = Query(..., description="Financial year in yyyy-yy format"),
    db: Session = Depends(get_db),
    current_investor: investor_models.Investor = Depends(get_current_investor),
):
    if "-" not in year:
        raise HTTPException(status_code=400, detail="Year format must be yyyy-yy, e.g., 2024-25")
    start_year, end_year = year.split("-")
    start_date = date(int(start_year), 4, 1)
    end_date = date(int("20" + end_year), 3, 31)

    calculate_and_store_capital_gains(db, current_investor.investor_id, start_date, end_date)

    return {"message": "Capital gains computed and stored successfully"}


@router.get("/investor/capital-gains-stored", response_model=List[investor_schemas.CapitalGainRecord])
def get_stored_capital_gains(
    year: str = Query(..., description="Financial year in yyyy-yy format"),
    db: Session = Depends(get_db),
    current_investor: investor_models.Investor = Depends(get_current_investor),
):
    if "-" not in year:
        raise HTTPException(status_code=400, detail="Year format must be yyyy-yy, e.g., 2024-25")
    start_year, end_year = year.split("-")
    start_date = date(int(start_year), 4, 1)
    end_date = date(int("20" + end_year), 3, 31)

    gains = (
        db.query(investor_models.CapitalGains)
        .filter(
            investor_models.CapitalGains.investor_id == current_investor.investor_id,
            investor_models.CapitalGains.sell_date >= start_date,
            investor_models.CapitalGains.sell_date <= end_date,
        )
        .order_by(investor_models.CapitalGains.sell_date.desc())
        .all()
    )

    return [
        investor_schemas.CapitalGainRecord(
            Scheme=g.scheme_id,
            Folio=g.folio_number,
            BuyDate=g.buy_date,
            SellDate=g.sell_date,
            BuyAmount=g.buy_amount,
            SellAmount=g.sell_amount,
            Gains=g.gains,
            Type=g.gain_type,
        )
        for g in gains
    ]

@router.get("/investor/valuation-report", response_model=List[investor_schemas.ValuationReportRecord])
def get_valuation_report(
    db: Session = Depends(get_db),
    current_investor: investor_models.Investor = Depends(get_current_investor),
):
    folios = (
        db.query(investor_models.FolioSummary)
        .filter(investor_models.FolioSummary.investor_id == current_investor.investor_id)
        .all()
    )

    response = [
        investor_schemas.ValuationReportRecord(
            scheme=folio.scheme_id,
            folio=folio.folio_number,
            units=folio.total_units,
            nav=folio.last_nav,
            value=folio.total_value,
        )
        for folio in folios
    ]

    return response


@router.get("/api/cas/generate-excel")
def generate_cas_excel(
    year: str = Query(..., description="Financial year yyyy-yy"),
    db: Session = Depends(get_db),
    current_investor=Depends(get_current_investor)
):
    if "-" not in year:
        raise HTTPException(status_code=400, detail="Year format must be yyyy-yy")

    start_year, end_year = year.split("-")
    try:
        start_date = date(int(start_year), 4, 1)
        end_date = date(int("20" + end_year), 3, 31)
    except Exception:
        raise HTTPException(status_code=400, detail="Invalid year format")

    try:
        txns = (
            db.query(TransactionLedger)
            .filter(
                TransactionLedger.investor_id == current_investor.investor_id,
                TransactionLedger.date >= start_date,
                TransactionLedger.date <= end_date,
            )
            .order_by(TransactionLedger.date)
            .all()
        )

        wb = Workbook()
        ws = wb.active
        ws.title = "CAS"

        headers = ["Date", "Transaction Type", "Scheme ID", "Folio Number", "Units", "Amount", "Status", "Bank", "Payment Mode"]
        ws.append(headers)

        for txn in txns:
            ws.append([
                txn.date.strftime("%Y-%m-%d"),
                txn.txn_type,
                txn.scheme_id,
                txn.folio_number,
                txn.units,
                txn.amount,
                txn.status,
                txn.bank,
                txn.payment_mode,
            ])

        # Adjust column widths
        for col_idx, col in enumerate(ws.columns, 1):
            max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
            ws.column_dimensions[get_column_letter(col_idx)].width = max_length + 2

        stream = io.BytesIO()
        wb.save(stream)
        stream.seek(0)

        headers = {
            "Content-Disposition": f"attachment; filename=CAS_{year}.xlsx"
        }
        stream = io.BytesIO()
        wb.save(stream)
        stream.seek(0)
        return Response(
            content=stream.read(),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers=headers,
        )

    except Exception as e:
        logging.exception("Failed to generate CAS Excel")
        raise HTTPException(status_code=500, detail="Failed to generate CAS Excel file")


def get_documents(current_investor=Depends(get_current_investor), db: Session = Depends(get_db)):
    docs = db.query(investor_models.Document).filter(investor_models.Document.investor_id == current_investor.investor_id).all()
    result = []
    for d in docs:
        result.append({
            "id": d.id, 
            "name": d.filename,
            "uploaded_on": d.upload_date.strftime("%Y-%m-%d"),
            "status": d.status
        })
    return {"documents": result}

@router.post("/documents")
async def upload_document(
    current_investor=Depends(get_current_investor),
    db: Session = Depends(get_db),
    file: UploadFile = File(...),
    document_type: str = 'supporting_document'
):
    import os, shutil
    from datetime import datetime

    upload_folder = "uploaded_documents"
    os.makedirs(upload_folder, exist_ok=True)
    filepath = f"{upload_folder}/{current_investor.investor_id}_{datetime.utcnow().timestamp()}_{file.filename}"

    with open(filepath, "wb") as buffer:
        shutil.copyfileobj(file.file, buffer)

    document = investor_models.Document(
        investor_id=current_investor.investor_id,
        filename=file.filename,
        path=filepath,
        upload_date=datetime.utcnow(),
        status="Pending",
        document_type=document_type
    )
    db.add(document)
    db.commit()
    db.refresh(document)

    return {
        "id": document.id,
        "name": document.filename,
        "uploaded_on": document.upload_date.strftime("%Y-%m-%d"),
        "status": document.status
    }

@router.post("/investor/documents")
def get_documents(
    current_investor=Depends(get_current_investor),
    db: Session = Depends(get_db)
):
    docs = db.query(investor_models.Document).filter(investor_models.Document.investor_id == current_investor.investor_id).all()
    result = []
    for d in docs:
        result.append({
            "id": d.id,
            "name": d.filename,
            "uploaded_on": d.upload_date.strftime("%Y-%m-%d"),
            "status": d.status
        })
    return {"documents": result}

@router.get("/investor/documents/download/{document_id}")
def download_document(
    document_id: int,
    current_investor=Depends(get_current_investor),
    db: Session = Depends(get_db)
):
    from fastapi.responses import FileResponse
    import os
    from fastapi import HTTPException

    document = db.query(investor_models.Document).filter(
        investor_models.Document.id == document_id,
        investor_models.Document.investor_id == current_investor.investor_id
    ).first()
    
    if not document:
        raise HTTPException(status_code=404, detail="Document not found")

    if not os.path.exists(document.path):
        raise HTTPException(status_code=404, detail="File not found on server")

    return FileResponse(
        path=document.path,
        filename=document.filename,
        media_type="application/octet-stream",
    )

@router.post("/investor/documents")
async def create_document(
    current_investor=Depends(get_current_investor),
    db: Session = Depends(get_db),
    file: UploadFile = File(...),
    document_type: str = 'supporting_document'
):
    # your upload logic here, save file, store in DB
    return {"message": "Document created successfully!"}

# GET route to retrieve uploaded documents for current user
@router.get("/documents")
def get_documents(
    current_investor=Depends(get_current_investor),
    db: Session = Depends(get_db)
):
    docs = db.query(investor_models.Document).filter(investor_models.Document.investor_id == current_investor.investor_id).all()
    result = []
    for doc in docs:
        result.append({
            "id": doc.id, 
            "name": doc.filename,
            "uploaded_on": doc.upload_date.strftime("%Y-%m-%d"),
            "status": doc.status
        })
    return {"documents": result}
@router.get("/mandates", response_model=List[investor_schemas.MandateResponse])
def get_mandates(current_investor=Depends(get_current_investor), db: Session = Depends(get_db)):
    mandates = db.query(investor_models.Mandate).filter(investor_models.Mandate.investor_id == current_investor.investor_id).all()
    return mandates

@router.post("/mandates", response_model=investor_schemas.MandateResponse, status_code=status.HTTP_201_CREATED)
def register_mandate(mandate: investor_schemas.MandateCreate, current_investor=Depends(get_current_investor), db: Session = Depends(get_db)):
    new_mandate = investor_models.Mandate(
        investor_id=current_investor.investor_id,
        scheme=mandate.scheme,
        type=mandate.type,
        status="Pending",
        created=mandate.created,
    )
    db.add(new_mandate)
    db.commit()
    db.refresh(new_mandate)
    return new_mandate


@router.delete("/mandates/{mandate_id}", status_code=status.HTTP_204_NO_CONTENT)
def revoke_mandate(mandate_id: int, current_investor=Depends(get_current_investor), db: Session = Depends(get_db)):
    mandate = db.query(investor_models.Mandate).filter(
        investor_models.Mandate.id == mandate_id, investor_models.Mandate.investor_id == current_investor.investor_id
    ).first()
    if mandate is None:
        raise HTTPException(status_code=404, detail="Mandate not found")
    mandate.status = "Revoked"
    db.commit()
    return Response(status_code=status.HTTP_204_NO_CONTENT)


@router.get("/service-requests", response_model=List[investor_schemas.ServiceRequestResponse])
def get_service_requests(current_investor=Depends(get_current_investor), db: Session = Depends(get_db)):
    requests = (
        db.query(investor_models.ServiceRequest)
        .filter(investor_models.ServiceRequest.investor_id == current_investor.investor_id)
        .all()
    )
    return requests


@router.post("/service-requests", response_model=investor_schemas.ServiceRequestResponse, status_code=status.HTTP_201_CREATED)
def raise_service_request(request: investor_schemas.ServiceRequestCreate, current_investor=Depends(get_current_investor), db: Session = Depends(get_db)):
    new_request = investor_models.ServiceRequest(
        investor_id=current_investor.investor_id,
        type=request.type,
        details=request.details,
        status="Open",
        created=request.created,
    )
    db.add(new_request)
    db.commit()
    db.refresh(new_request)
    return new_request

@router.get("/notifications", response_model=List[investor_schemas.NotificationResponse])
def get_notifications(current_investor=Depends(get_current_investor), db: Session = Depends(get_db)):
    notifications = (
        db.query(investor_models.Notification)
        .filter(investor_models.Notification.investor_id == current_investor.investor_id)
        .order_by(investor_models.Notification.date.desc())
        .all()
    )
    return notifications

@router.post("/notifications/{notification_id}/mark-read", status_code=status.HTTP_200_OK)
def mark_notification_read(notification_id: int, current_investor=Depends(get_current_investor), db: Session = Depends(get_db)):
    notification = (
        db.query(investor_models.Notification)
        .filter(investor_models.Notification.id == notification_id, investor_models.Notification.investor_id == current_investor.investor_id)
        .first()
    )
    if not notification:
        raise HTTPException(status_code=404, detail="Notification not found")
    notification.read = True
    db.commit()
    return {"message": "Notification marked as read"}

@router.post("/notifications/clear-all", status_code=status.HTTP_200_OK)
def clear_all_notifications(current_investor=Depends(get_current_investor), db: Session = Depends(get_db)):
    db.query(investor_models.Notification) \
        .filter(investor_models.Notification.investor_id == current_investor.investor_id) \
        .update({"read": True})
    db.commit()
    return {"message": "All notifications marked as read"}


@router.get("/complaints", response_model=List[investor_schemas.ComplaintResponse])
def get_complaints(current_investor=Depends(get_current_investor), db: Session = Depends(get_db)):
    complaints = (
        db.query(investor_models.Complaint)
        .filter(investor_models.Complaint.investor_id == current_investor.investor_id)
        .order_by(investor_models.Complaint.date.desc())
        .all()
    )
    return complaints

@router.post("/complaints", response_model=investor_schemas.ComplaintResponse, status_code=status.HTTP_201_CREATED)
def create_complaint(complaint: investor_schemas.ComplaintCreate, current_investor=Depends(get_current_investor), db: Session = Depends(get_db)):
    new_complaint = investor_models.Complaint(
        investor_id=current_investor.investor_id,
        subject=complaint.subject,
        description=complaint.description,
        status=complaint.status,
        date=complaint.date,
    )
    db.add(new_complaint)
    db.commit()
    db.refresh(new_complaint)
    return new_complaint

@router.get("/support-tickets", response_model=List[investor_schemas.SupportTicketResponse])
def get_support_tickets(current_investor=Depends(get_current_investor), db: Session = Depends(get_db)):
    tickets = (
        db.query(investor_models.SupportTicket)
        .filter(investor_models.SupportTicket.investor_id == current_investor.investor_id)
        .order_by(investor_models.SupportTicket.created.desc())
        .all()
    )
    return tickets

@router.post("/support-tickets", response_model=investor_schemas.SupportTicketResponse, status_code=status.HTTP_201_CREATED)
def create_support_ticket(ticket: investor_schemas.SupportTicketCreate, current_investor=Depends(get_current_investor), db: Session = Depends(get_db)):
    new_ticket = investor_models.SupportTicket(
        investor_id=current_investor.investor_id,
        subject=ticket.subject,
        message=ticket.message,
        status=ticket.status,
        created=ticket.created,
    )
    db.add(new_ticket)
    db.commit()
    db.refresh(new_ticket)
    return new_ticket

@router.get("/clients", response_model=List[investor_schemas.ClientResponse])
def get_clients(current_investor=Depends(get_current_investor), db: Session = Depends(get_db)):
    # You may filter clients related to current_investor if needed
    clients = db.query(investor_models.Client).all()
    return clients

@router.get("/transactions", response_model=List[investor_schemas.TransactionResponse])
def get_transactions(
    current_investor: investor_models.Investor = Depends(get_current_investor),
    db: Session = Depends(get_db),
):
    transactions = (
        db.query(investor_models.Transaction)
        .filter(investor_models.Transaction.investor_id == current_investor.investor_id)
        .order_by(investor_models.Transaction.date.desc())
        .limit(30)
        .all()
    )

    result = []
    for txn in transactions:
        result.append(
            investor_schemas.TransactionResponse(
                txn_id=txn.txn_id,
                folio_number=txn.folio_number or "",
                date=txn.date,
                scheme_id=txn.scheme_id,
                txn_type=txn.txn_type,
                units=txn.units,
                nav=txn.nav,
                amount=txn.amount,
                plan=txn.plan,
                status=txn.status,
                description=f"{txn.txn_type} transaction",
            )
        )
    return result


@router.get("/folios/{folio_number}/holdings", response_model=List[investor_schemas.ValuationReportRecord])
def get_folio_holdings(
    folio_number: str,
    db: Session = Depends(get_db),
    current_investor=Depends(get_current_investor),
):
    holdings = db.query(investor_models.FolioSummary).filter(
        investor_models.FolioSummary.folio_number == folio_number,
        investor_models.FolioSummary.investor_id == current_investor.investor_id,
    ).all()

    return [
        investor_schemas.ValuationRecord(
            scheme=h.scheme_id,
            folio=h.folio_number,
            units=h.units,
            nav=h.nav,
            value=h.value,
        )
        for h in holdings
    ]


@router.get("/folios/{folio_number}/transactions", response_model=List[investor_schemas.TransactionResponse])
def get_folio_transactions(
    folio_number: str,
    db: Session = Depends(get_db),
    current_investor=Depends(get_current_investor),
):
    txns = (
        db.query(investor_models.TransactionLedger)
        .filter(
            investor_models.TransactionLedger.folio_number == folio_number,
            investor_models.TransactionLedger.investor_id == current_investor.investor_id,
        )
        .order_by(investor_models.TransactionLedger.date.desc())
        .limit(30)
        .all()
    )

    return [
        investor_schemas.TransactionResponse(
            txn_id=txn.txn_id,
            folio_number=txn.folio_number or "",
            date=txn.date,
            scheme_id=txn.scheme_id,
            txn_type=txn.txn_type,
            units=txn.units,
            nav=txn.nav,
            amount=txn.amount,
            plan=txn.plan,
            status=txn.status,
            description=f"{txn.txn_type} transaction",
        )
        for txn in txns
    ]


@router.get("/investor/transactionhistory", response_model=List[TransactionLedgerResponse])
def get_transaction_history(
    db: Session = Depends(get_db),
    current_investor=Depends(get_current_investor),
):
    transactions = (
        db.query(investor_models.TransactionLedger)
        .filter(investor_models.TransactionLedger.investor_id == current_investor.investor_id)
        .order_by(investor_models.TransactionLedger.date.desc())
        .all()
    )
    return transactions
